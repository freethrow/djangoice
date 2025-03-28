import openpyxl
import csv
import os
from datetime import datetime

def determine_book(row_number):
    """Determine which book the test belongs to based on row number."""
    if 8 <= row_number <= 13:
        return "1. knjiga testova"
    elif 14 <= row_number <= 17:
        return "2. knjiga testova"
    elif 18 <= row_number <= 21:
        return "3. knjiga testova"
    elif 22 <= row_number <= 25:
        return "4. knjiga testova"
    elif 26 <= row_number <= 29:
        return "5. knjiga testova"
    elif 30 <= row_number <= 33:
        return "6. knjiga testova"
    else:
        return "Unknown book"

def get_skill_from_column(col_index):
    """Map column index to skill name."""
    if 3 <= col_index <= 5:  # D-F (0-indexed: 3-5)
        return "Use of English"
    elif 6 <= col_index <= 8:  # G-I
        return "Reading"
    elif 9 <= col_index <= 11:  # J-L
        return "Writing"
    elif 12 <= col_index <= 14:  # M-O
        return "Listening"
    elif 15 <= col_index <= 17:  # P-R
        return "Speaking"
    return None

# Load the workbook
print("Loading Excel file...")
workbook = openpyxl.load_workbook('student_data.xlsx')

# Create a list to store all student test data
all_student_data = []
flattened_data = []  # This will store the data in a format suitable for CSV

# Process each sheet in the workbook
print(f"Processing {len(workbook.worksheets)} sheets...")
for sheet_index, sheet in enumerate(workbook.worksheets, 1):
    sheet_name = sheet.title
    print(f"Processing sheet {sheet_index}/{len(workbook.worksheets)}: {sheet_name}")
    
    # Iterate through all rows that might contain test data
    for row_number in range(1, sheet.max_row + 1):
        # Get the value in column C for this row
        test_number_cell = sheet.cell(row=row_number, column=3).value
        
        # Check if C column has a test number (integer)
        try:
            if test_number_cell is not None and isinstance(test_number_cell, (int, float)) and test_number_cell == int(test_number_cell):
                test_number = int(test_number_cell)
                book = determine_book(row_number)
                
                # Check the result columns (D, G, J, M, P)
                result_columns = [3, 6, 9, 12, 15]  # 0-indexed: D, G, J, M, P
                
                for col_index in result_columns:
                    result_value = sheet.cell(row=row_number, column=col_index + 1).value
                    
                    # Check if there's an integer in the result column
                    if result_value is not None and isinstance(result_value, (int, float)) and result_value == int(result_value):
                        # Get the max value from the adjacent column
                        max_value = sheet.cell(row=row_number, column=col_index + 2).value
                        
                        # Get the skill name based on column
                        skill_name = get_skill_from_column(col_index)
                        
                        # Add skill data if both result and max are available
                        if max_value is not None and isinstance(max_value, (int, float)):
                            # Create a flattened record for CSV
                            csv_record = {
                                "student_name": sheet_name,
                                "book": book,
                                "test_number": test_number,
                                "skill": skill_name,
                                "result": int(result_value),
                                "max": int(max_value) if max_value == int(max_value) else float(max_value),
                                "percentage": round(int(result_value) / float(max_value) * 100, 2)
                            }
                            flattened_data.append(csv_record)
        except (ValueError, TypeError):
            # Skip if not an integer
            continue

# Create a timestamped filename for the CSV
timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
csv_filename = f"student_test_results_{timestamp}.csv"

# Define the field names for the CSV
fieldnames = ["student_name", "book", "test_number", "skill", "result", "max", "percentage"]

# Write the data to a CSV file
print(f"Writing {len(flattened_data)} records to {csv_filename}...")
with open(csv_filename, 'w', newline='', encoding='utf-8') as csvfile:
    writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
    writer.writeheader()
    writer.writerows(flattened_data)

print(f"Data successfully exported to {csv_filename}")
print(f"Total records: {len(flattened_data)}")

# Print some basic statistics
if flattened_data:
    total_students = len(set(record["student_name"] for record in flattened_data))
    total_tests = len(set((record["student_name"], record["book"], record["test_number"]) for record in flattened_data))
    avg_percentage = sum(record["percentage"] for record in flattened_data) / len(flattened_data)
    
    print(f"\nSummary Statistics:")
    print(f"Total students: {total_students}")
    print(f"Total unique tests: {total_tests}")
    print(f"Average score percentage: {avg_percentage:.2f}%")