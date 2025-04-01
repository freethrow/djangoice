# eventi/views.py

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from django.core.files.base import ContentFile

from django.shortcuts import render, redirect
from django.urls import reverse_lazy
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib import messages
from django.views.generic import (
    ListView,
    DetailView,
    CreateView,
    UpdateView,
    DeleteView,
    TemplateView,
)
from django.db.models import Q
from django.db.models import Case, When, Value, IntegerField

from django.utils.translation import gettext_lazy as _

from django.views.generic import FormView
from django.urls import reverse

from django.http import FileResponse, Http404

from django.shortcuts import get_object_or_404

from django.http import HttpResponse
from django import forms

# import settings
from django.conf import settings

import datetime
import boto3


from .models import Event
from django.core.serializers import serialize
import json

import os
import tempfile
from docxtpl import DocxTemplate

from django.views import View
from django.core.files.storage import default_storage
from storages.backends.s3boto3 import S3Boto3Storage

from botocore.client import Config


from django.views.decorators.csrf import csrf_exempt
from django.utils.decorators import method_decorator

from .models import EventFile
from .forms import EventForm


# Public views (read-only)


class EventListView(ListView):
    model = Event
    template_name = "eventi/event_list.html"
    context_object_name = "events"
    paginate_by = 10

    def get_queryset(self):
        # Start with a queryset that includes related models to avoid N+1 queries
        queryset = Event.objects.all().select_related("settore")

        # Handle search and filters
        q = self.request.GET.get("q")
        categoria = self.request.GET.get("categoria")
        office = self.request.GET.get("office")
        paese = self.request.GET.get("paese")
        settore = self.request.GET.get("settore")

        # Build query conditions incrementally instead of chaining filters
        conditions = Q()
        if q:
            conditions &= Q(titolo__icontains=q)
        if categoria:
            conditions &= Q(categoria=categoria)
        if office:
            conditions &= Q(office=office)
        if paese:
            conditions &= Q(paese=paese)
        if settore:
            conditions &= Q(settore_id=settore)

        # Apply all filters at once
        if conditions:
            queryset = queryset.filter(conditions)

        # Handle sorting
        sort_field = self.request.GET.get("sort", "data_inizio")
        sort_direction = self.request.GET.get("direction", "desc")

        # Validate sort field to prevent SQL injection
        allowed_fields = [
            "titolo",
            "data_inizio",
            "citta",
            "settore",
            "tipologia",
            "office",
        ]
        if sort_field not in allowed_fields:
            sort_field = "data_inizio"

        # Special handling for settore field (now a foreign key)
        if sort_field == "settore":
            # Use joined field instead of lookup
            if sort_direction == "asc":
                queryset = queryset.order_by("settore__nome")
            else:
                queryset = queryset.order_by("-settore__nome")
        # Special handling for office field (Belgrado first)
        elif sort_field == "office":
            # Use Case/When but with a single annotation
            queryset = queryset.annotate(
                office_order=Case(
                    When(office="Belgrado", then=Value(0)),
                    default=Value(1),
                    output_field=IntegerField(),
                )
            )
            if sort_direction == "asc":
                queryset = queryset.order_by("office_order", "office")
            else:
                queryset = queryset.order_by("-office_order", "-office")
        else:
            # Apply sorting direction for other fields
            if sort_direction == "asc":
                order_by = sort_field
            else:
                order_by = f"-{sort_field}"
            queryset = queryset.order_by(order_by)

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Cache frequently accessed data
        from django.core.cache import cache

        # Try to get categories from cache first
        categoria_choices = cache.get("categoria_choices")
        if not categoria_choices:
            categoria_choices = Event.CATEGORIA_CHOICES
            # Cache for 1 hour (3600 seconds)
            cache.set("categoria_choices", categoria_choices, 3600)
        context["categoria_choices"] = categoria_choices

        # Same for other static choices
        context["office_choices"] = Event.OFFICE_CHOICES
        context["paese_choices"] = Event.COUNTRY_CHOICES

        # For settore choices, use a cached query
        settore_choices = cache.get("settore_choices")
        if not settore_choices:
            from .models import Settore

            settore_choices = [
                (s.id, s.nome) for s in Settore.objects.all().order_by("nome")
            ]
            cache.set("settore_choices", settore_choices, 3600)
        context["settore_choices"] = settore_choices

        # Add current filter values for re-selecting in dropdowns
        context["current_categoria"] = self.request.GET.get("categoria", "")
        context["current_office"] = self.request.GET.get("office", "")
        context["current_paese"] = self.request.GET.get("paese", "")
        context["current_settore"] = self.request.GET.get("settore", "")

        return context


class EventDetailView(DetailView):
    model = Event
    template_name = "eventi/event_detail.html"
    context_object_name = "event"

    def get_object(self, queryset=None):
        obj = super().get_object(queryset)
        # Allow view if: event is public, user is creator, or user is staff
        if obj.public or (
            self.request.user.is_authenticated
            and (obj.created_by == self.request.user or self.request.user.is_staff)
        ):
            return obj
        # Otherwise redirect to the events list
        messages.error(
            self.request, "Non hai l'autorizzazione per visualizzare questo evento."
        )
        return redirect("event_list")


class EventCreateView(LoginRequiredMixin, CreateView):
    model = Event
    form_class = EventForm
    template_name = "eventi/event_form.html"
    success_url = reverse_lazy("event_list")

    def form_valid(self, form):
        # Set the created_by and last_updated_by fields
        form.instance.created_by = self.request.user
        form.instance.last_updated_by = self.request.user

        # First save the event to get an ID
        response = super().form_valid(form)

        # Handle the single file upload
        if "file" in self.request.FILES:
            file = self.request.FILES["file"]
            EventFile.objects.create(
                event=self.object,
                file=file,
                title=file.name,  # Use filename as default title
                created_by=self.request.user,
            )

        messages.success(self.request, "Evento creato con successo.")
        return response


class EventUpdateView(LoginRequiredMixin, UpdateView):
    model = Event
    form_class = EventForm
    template_name = "eventi/event_form.html"
    success_url = reverse_lazy("event_list")

    def get_queryset(self):
        if self.request.user.is_staff:
            return Event.objects.all()
        return Event.objects.filter(created_by=self.request.user)

    def form_valid(self, form):
        form.instance.last_updated_by = self.request.user

        response = super().form_valid(form)

        # Handle the single file upload
        if "file" in self.request.FILES:
            file = self.request.FILES["file"]
            EventFile.objects.create(
                event=self.object,
                file=file,
                title=file.name,  # Use filename as default title
                created_by=self.request.user,
            )

        messages.success(self.request, "Evento aggiornato con successo.")
        return response

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Add existing files to the context
        if self.object:
            context["event_files"] = EventFile.objects.filter(event=self.object)
        return context


def get_upload_path(instance, filename):
    """Generate a path for file uploads based on event and date"""
    return f"event_files/{instance.event.id}/{filename}"


class EventFileUploadView(LoginRequiredMixin, FormView):
    template_name = "eventi/eventfile_form.html"

    def get_form_class(self):
        # Create a dynamic form class
        class FileUploadForm(forms.Form):
            file = forms.FileField(
                label=_("Documento"),
                widget=forms.ClearableFileInput(attrs={"class": "block w-full"}),
            )
            title = forms.CharField(
                label=_("Titolo"),
                required=False,
                widget=forms.TextInput(attrs={"class": "block w-full"}),
            )
            description = forms.CharField(
                label=_("Descrizione"),
                required=False,
                widget=forms.Textarea(attrs={"rows": 2, "class": "block w-full"}),
            )

        return FileUploadForm

    def form_valid(self, form):
        event_id = self.kwargs.get("pk")
        event = get_object_or_404(Event, pk=event_id)

        # Check permissions
        if not self.request.user.is_staff and event.created_by != self.request.user:
            messages.error(
                self.request,
                _("Non hai i permessi per aggiungere file a questo evento."),
            )
            return redirect("event_detail", pk=event_id)

        try:
            # Create the file
            file = self.request.FILES["file"]
            title = (
                form.cleaned_data.get("title")
                or os.path.splitext(os.path.basename(file.name))[0]
            )
            description = form.cleaned_data.get("description", "")

            EventFile.objects.create(
                event=event,
                file=file,
                title=title,
                description=description,
                created_by=self.request.user,
            )

            messages.success(self.request, _("File aggiunto con successo."))
        except Exception as e:
            messages.error(
                self.request,
                _(
                    "Si è verificato un errore durante il caricamento del file: {}"
                ).format(str(e)),
            )

        return redirect("event_detail", pk=event_id)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        event_id = self.kwargs.get("pk")
        context["event"] = get_object_or_404(Event, pk=event_id)
        return context


class EventFileDeleteView(LoginRequiredMixin, DeleteView):
    model = EventFile
    template_name = "eventi/eventfile_confirm_delete.html"
    pk_url_kwarg = "file_pk"

    def get_success_url(self):
        # Redirect back to the event detail page after deletion
        return reverse("event_detail", kwargs={"pk": self.object.event.id})

    def get_object(self, queryset=None):
        # Get the file, filtering by event_pk for added security
        return get_object_or_404(
            EventFile,
            pk=self.kwargs.get("file_pk"),
            event__pk=self.kwargs.get("event_pk"),
        )

    def delete(self, request, *args, **kwargs):
        self.object = self.get_object()
        event = self.object.event

        # Check permissions
        if not (request.user.is_staff or event.created_by == request.user):
            messages.error(request, "Non hai i permessi per eliminare questo file.")
            return redirect("event_detail", pk=event.pk)

        # Delete the physical file from storage first
        try:
            if self.object.file and default_storage.exists(self.object.file.name):
                default_storage.delete(self.object.file.name)
        except Exception as e:
            # Log the error, but don't stop the deletion process
            messages.warning(
                request, f"Errore durante l'eliminazione del file dal storage: {str(e)}"
            )

        # Perform the database record deletion and add a success message
        response = super().delete(request, *args, **kwargs)
        messages.success(request, "File eliminato con successo.")

        return response


# Function-based view alternative for event creation
@login_required
def create_event(request):
    if request.method == "POST":
        form = EventForm(request.POST)
        if form.is_valid():
            event = form.save(commit=False)
            event.created_by = request.user
            event.last_updated_by = request.user
            event.save()
            messages.success(request, "Evento creato con successo.")
            return redirect("event_list")
    else:
        form = EventForm()

    return render(request, "eventi/event_form.html", {"form": form})


class EventDeleteView(LoginRequiredMixin, DeleteView):
    model = Event
    template_name = "eventi/event_confirm_delete.html"
    success_url = reverse_lazy("event_list")

    def get_queryset(self):
        # Staff can delete any event
        if self.request.user.is_staff:
            return Event.objects.all()
        # Regular users can only delete their own events
        return Event.objects.filter(created_by=self.request.user)

    def delete(self, request, *args, **kwargs):
        messages.success(request, "Evento eliminato con successo.")
        return super().delete(request, *args, **kwargs)


def public_events_api(request):
    """
    API endpoint to list all public events.
    Returns a JSON response with all events that have public=True.
    No authentication required.
    """
    # Query all public events
    events = Event.objects.filter(public=True)

    # Serialize the queryset to JSON
    events_json = serialize("json", events)
    events_data = json.loads(events_json)

    # Format the response to include only relevant fields and flatten the structure
    formatted_events = []
    for event in events_data:
        event_fields = event["fields"]
        formatted_event = {
            "id": event["pk"],
            "categoria": event_fields["categoria"],
            "office": event_fields["office"],
            "titolo": event_fields["titolo"],
            "data_inizio": event_fields["data_inizio"],
            "data_fine": event_fields["data_fine"],
            "paese": event_fields["paese"],
            "citta": event_fields["citta"],
            "settore": event_fields["settore"],
            "tipologia": event_fields["tipologia"],
            "descrizione": event_fields["descrizione"],
            "created_at": event_fields["created_at"],
            "updated_at": event_fields["updated_at"],
        }
        formatted_events.append(formatted_event)

    # Return the formatted events as JSON response
    return JsonResponse({"events": formatted_events}, safe=True)


class EventFileDownloadView(LoginRequiredMixin, View):
    def get(self, request, event_pk, file_pk):
        event_file = get_object_or_404(EventFile, pk=file_pk, event__pk=event_pk)

        try:
            response = FileResponse(event_file.file.open("rb"))
            filename = event_file.file.name.split("/")[-1]
            response["Content-Disposition"] = f'attachment; filename="{filename}"'
            return response
        except Exception:
            raise Http404("Il file richiesto non è disponibile.")


class ReportSelectionView(LoginRequiredMixin, TemplateView):
    template_name = "eventi/report_selection.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Get events filtered by user permissions
        if self.request.user.is_staff:
            events = Event.objects.all()
        else:
            events = Event.objects.filter(public=True) | Event.objects.filter(
                created_by=self.request.user
            )

        # Apply filters if provided
        categoria = self.request.GET.get("categoria")
        if categoria:
            events = events.filter(categoria=categoria)

        paese = self.request.GET.get("paese")
        if paese:
            events = events.filter(paese=paese)

        year = self.request.GET.get("year")
        if year:
            events = events.filter(data_inizio__year=year)

        # Sort events by date
        events = events.order_by("data_inizio")

        # Get unique filter options for dropdowns
        context["events"] = events

        context["paesi"] = Event.COUNTRY_CHOICES

        # Get unique years from events
        years = events.dates("data_inizio", "year")
        context["years"] = [date.year for date in years]

        # Active filters
        context["active_filters"] = {
            "categoria": categoria,
            "paese": paese,
            "year": year,
        }

        return context


@method_decorator(csrf_exempt, name="dispatch")
class GenerateReportView(LoginRequiredMixin, TemplateView):
    template_name = "eventi/report_result.html"

    def post(self, request, *args, **kwargs):
        # Get selected event IDs
        event_ids = request.POST.getlist("event_ids")
        export_format = request.POST.get("export_format", "docx")

        if not event_ids:
            messages.error(request, "Seleziona almeno un evento da esportare.")
            return redirect("report_selection")

        # Get events
        events = Event.objects.filter(id__in=event_ids).order_by("data_inizio")

        if export_format == "docx":
            return self.generate_docx(events)
        elif export_format == "excel":  # Changed from 'pdf' to 'excel'
            return self.generate_excel(events)
        else:
            messages.error(request, "Formato non supportato.")
            return redirect("report_selection")

    def generate_docx(self, events):
        # Create a context for the template
        context = {
            "events": [
                {
                    "titolo": event.titolo,
                    "categoria": event.categoria,
                    "data_inizio": event.data_inizio.strftime("%d/%m/%Y"),
                    "data_fine": (
                        event.data_fine.strftime("%d/%m/%Y") if event.data_fine else ""
                    ),
                    "paese": event.paese,
                    "citta": event.citta,
                    "settore": event.settore,
                    "tipologia": event.tipologia,
                    "descrizione": event.descrizione,
                    "created_by": str(event.created_by) if event.created_by else "",
                }
                for event in events
            ],
            "total_events": len(events),
            "report_date": datetime.datetime.now().strftime("%d/%m/%Y %H:%M"),
        }

        # Path to your Word template
        template_path = os.path.join(
            settings.BASE_DIR, "eventi", "templates", "eventi", "report_template.docx"
        )

        # Check if template exists
        if not os.path.exists(template_path):
            messages.error(self.request, "Template di report non trovato.")
            return redirect("report_selection")

        # Use a context manager to ensure file is properly closed
        try:
            # Create a unique filename to reduce conflict chances
            unique_filename = f"eventi_report_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.docx"
            output_path = os.path.join(tempfile.gettempdir(), unique_filename)

            # Generate the document
            doc = DocxTemplate(template_path)
            doc.render(context)
            doc.save(output_path)

            # Open and read the file in binary mode
            with open(output_path, "rb") as f:
                file_content = f.read()

            # Prepare HTTP response
            response = HttpResponse(
                file_content,
                content_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
            )
            response["Content-Disposition"] = f"attachment; filename={unique_filename}"

            # Try to remove the temporary file
            try:
                os.unlink(output_path)
            except Exception as cleanup_error:
                # Log the cleanup error but don't interrupt the response
                print(
                    f"Warning: Could not delete temporary file {output_path}. Error: {cleanup_error}"
                )

            return response

        except Exception as e:
            # Comprehensive error handling
            error_message = f"Errore nella generazione del report: {str(e)}"
            messages.error(self.request, error_message)

            # Try to clean up any temporary file that might have been created
            if "output_path" in locals() and os.path.exists(output_path):
                try:
                    os.unlink(output_path)
                except Exception as cleanup_error:
                    print(f"Additional cleanup error: {cleanup_error}")

            return redirect("report_selection")

    def generate_excel(self, events):
        """Generate Excel report with event data"""
        # Create a new workbook and select the active worksheet
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Eventi Report"

        # Define column headers based on model fields
        headers = [
            "ID",
            "Titolo",
            "Categoria",
            "Data Inizio",
            "Data Fine",
            "Paese",
            "Città",
            "Settore",
            "Tipologia",
            "Descrizione",
            "Pubblico",
            "Creato Da",
            "Data Creazione",
            "Ultimo Aggiornamento",
        ]

        # Apply header formatting
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(
            start_color="4682B4", end_color="4682B4", fill_type="solid"
        )

        # Write headers
        for col_num, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # Write data rows
        for row_num, event in enumerate(events, 2):
            # Map event attributes to columns
            row_data = [
                event.id,
                event.titolo,
                event.get_categoria_display(),  # Use display value for choices
                event.data_inizio.strftime("%d/%m/%Y"),
                event.data_fine.strftime("%d/%m/%Y") if event.data_fine else "",
                event.get_paese_display(),  # Use display value for choices
                event.citta,
                str(event.settore) if event.settore else "",
                event.tipologia,
                event.descrizione,
                "Sì" if event.public else "No",
                str(event.created_by) if event.created_by else "",
                event.created_at.strftime("%d/%m/%Y %H:%M") if event.created_at else "",
                event.updated_at.strftime("%d/%m/%Y %H:%M") if event.updated_at else "",
            ]

            # Write each cell in the row
            for col_num, cell_value in enumerate(row_data, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = cell_value

                # For description column, set wrap text
                if col_num == headers.index("Descrizione") + 1:
                    cell.alignment = Alignment(wrap_text=True)

        # Auto-adjust column widths
        for col_num, header in enumerate(headers, 1):
            # Set a minimum width for each column
            min_width = 10
            max_width = 40

            # Find the maximum content length in the column
            max_length = len(str(header))
            for row_num in range(2, len(events) + 2):
                cell_value = worksheet.cell(row=row_num, column=col_num).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))

            # Calculate adjusted width, constrained between min and max
            adjusted_width = min(max(min_width, max_length + 2), max_width)
            worksheet.column_dimensions[
                get_column_letter(col_num)
            ].width = adjusted_width

            # For description column, always use max width
            if header == "Descrizione":
                worksheet.column_dimensions[
                    get_column_letter(col_num)
                ].width = max_width

        # Generate timestamp for filename
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"report_{timestamp}.xlsx"

        # Create a temporary file to save the workbook
        output_path = os.path.join(tempfile.gettempdir(), filename)

        try:
            # Save workbook to temporary file
            workbook.save(output_path)

            # Create options for file storage
            option_save_to_storage = (
                self.request.POST.get("save_to_storage", "false") == "true"
            )

            # If requested to save to storage AND download
            if option_save_to_storage:
                # Open the file for reading in binary mode
                with open(output_path, "rb") as f:
                    file_content = f.read()

                # Save to Backblaze S3 storage
                storage_path = f"reports/{filename}"

                # Use default_storage to save the file (which uses your S3 configuration)
                default_storage.save(storage_path, ContentFile(file_content))

                # Add success message
                messages.success(
                    self.request,
                    f"Report salvato con successo nel cloud storage come '{storage_path}'",
                )

            # Return the file as a downloadable response
            with open(output_path, "rb") as f:
                file_content = f.read()

            # Prepare HTTP response
            response = HttpResponse(
                file_content,
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            response["Content-Disposition"] = f"attachment; filename={filename}"

            # Try to remove the temporary file
            try:
                os.unlink(output_path)
            except Exception as cleanup_error:
                print(
                    f"Warning: Could not delete temporary file {output_path}. Error: {cleanup_error}"
                )

            return response

        except Exception as e:
            # Error handling
            error_message = f"Errore nella generazione del report Excel: {str(e)}"
            messages.error(self.request, error_message)

            # Clean up any temporary file that might have been created
            if os.path.exists(output_path):
                try:
                    os.unlink(output_path)
                except Exception:
                    pass

            return redirect("report_selection")


def check_storage(request):
    """Debug view to check storage configuration"""
    storage_class = default_storage.__class__.__name__
    is_s3 = isinstance(default_storage, S3Boto3Storage)
    bucket_name = getattr(default_storage, "bucket_name", "Not set")
    endpoint_url = getattr(default_storage, "endpoint_url", "Not set")

    storage_info = f"""
    <h1>Storage Configuration</h1>
    <p><strong>Storage Class:</strong> {storage_class}</p>
    <p><strong>Is S3Boto3Storage:</strong> {is_s3}</p>
    <p><strong>Bucket Name:</strong> {bucket_name}</p>
    <p><strong>Endpoint URL:</strong> {endpoint_url}</p>
    """

    return HttpResponse(storage_info)


class ReportFileListView(LoginRequiredMixin, ListView):
    """View to display all reports stored in the Backblaze storage"""

    template_name = "eventi/report_files.html"
    context_object_name = "report_files"
    paginate_by = 20

    def get_queryset(self):
        # List files in the reports directory
        try:
            s3 = boto3.client(
                "s3",
                endpoint_url=settings.AWS_S3_ENDPOINT_URL,
                aws_access_key_id=settings.AWS_ACCESS_KEY_ID,
                aws_secret_access_key=settings.AWS_SECRET_ACCESS_KEY,
                region_name=settings.AWS_S3_REGION_NAME,
                config=Config(
                    signature_version="s3v4",
                    s3={"payload_signing_enabled": False},
                    # Remove the checksums parameter
                ),
            )

            response = s3.list_objects_v2(
                Bucket=settings.AWS_STORAGE_BUCKET_NAME, Prefix="reports/"
            )

            # If there are no files
            if "Contents" not in response:
                return []

            # Create list of file information
            files = []
            for item in response["Contents"]:
                # Skip the directory itself
                if item["Key"] == "reports/":
                    continue

                file_info = {
                    "key": item["Key"],
                    "name": os.path.basename(item["Key"]),
                    "size": item["Size"],
                    "last_modified": item["LastModified"],
                }
                files.append(file_info)

            # Sort by last modified (newest first)
            files.sort(key=lambda x: x["last_modified"], reverse=True)
            return files

        except Exception as e:
            # Log the error but return an empty list
            print(f"Error listing report files: {str(e)}")
            messages.error(self.request, f"Errore nel recupero dei file: {str(e)}")
            return []

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Convert file sizes to human-readable format
        for file in context["report_files"]:
            file["size_human"] = self.human_readable_size(file["size"])

        return context

    def human_readable_size(self, size, decimal_places=2):
        """Convert bytes to human-readable size"""
        for unit in ["B", "KB", "MB", "GB", "TB"]:
            if size < 1024.0:
                break
            size /= 1024.0
        return f"{size:.{decimal_places}f} {unit}"


class ReportFileDownloadView(LoginRequiredMixin, View):
    """View to download a report file from storage"""

    def get(self, request, *args, **kwargs):
        file_key = request.GET.get("key")

        if not file_key:
            messages.error(request, "Nessun file specificato.")
            return redirect("report_files")

        try:
            # Ensure the file key starts with 'reports/'
            if not file_key.startswith("reports/"):
                file_key = f"reports/{file_key}"

            # Get the file from S3 storage
            s3 = boto3.client(
                "s3",
                endpoint_url=settings.AWS_S3_ENDPOINT_URL,
                aws_access_key_id=settings.AWS_ACCESS_KEY_ID,
                aws_secret_access_key=settings.AWS_SECRET_ACCESS_KEY,
                region_name=settings.AWS_S3_REGION_NAME,
                config=Config(
                    signature_version="s3v4",
                    s3={"payload_signing_enabled": False},
                    # Remove the checksums parameter
                ),
            )

            response = s3.get_object(
                Bucket=settings.AWS_STORAGE_BUCKET_NAME, Key=file_key
            )

            # Determine content type based on file extension
            filename = os.path.basename(file_key)
            content_type = None

            if filename.endswith(".xlsx"):
                content_type = (
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            elif filename.endswith(".docx"):
                content_type = "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
            else:
                content_type = "application/octet-stream"

            # Create HTTP response with file content
            http_response = HttpResponse(
                response["Body"].read(), content_type=content_type
            )
            http_response["Content-Disposition"] = f'attachment; filename="{filename}"'

            return http_response

        except Exception as e:
            messages.error(request, f"Errore nel download del file: {str(e)}")
            return redirect("report_files")


class ReportFileDeleteView(LoginRequiredMixin, View):
    """View to delete a report file from storage"""

    def post(self, request, *args, **kwargs):
        file_key = request.POST.get("key")

        if not file_key:
            messages.error(request, "Nessun file specificato.")
            return redirect("report_files")

        try:
            # Ensure the file key starts with 'reports/'
            if not file_key.startswith("reports/"):
                file_key = f"reports/{file_key}"

            # Delete the file from S3 storage
            s3 = boto3.client(
                "s3",
                endpoint_url=settings.AWS_S3_ENDPOINT_URL,
                aws_access_key_id=settings.AWS_ACCESS_KEY_ID,
                aws_secret_access_key=settings.AWS_SECRET_ACCESS_KEY,
                region_name=settings.AWS_S3_REGION_NAME,
                config=Config(
                    signature_version="s3v4",
                    s3={"payload_signing_enabled": False},
                    # Remove the checksums parameter
                ),
            )

            s3.delete_object(Bucket=settings.AWS_STORAGE_BUCKET_NAME, Key=file_key)

            messages.success(
                request, f"File '{os.path.basename(file_key)}' eliminato con successo."
            )

        except Exception as e:
            messages.error(request, f"Errore nell'eliminazione del file: {str(e)}")

        return redirect("report_files")
